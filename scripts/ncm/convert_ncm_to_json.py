#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor de Planilha NCM para JSON
Converte planilha Excel simplificada (sem cabeçalho) para JSON estruturado
com hierarquia: Level -> Capítulo -> NCMs
"""

import json
import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# Caminhos padrão (relativos à raiz do projeto Chat UI)
_PROJECT_ROOT = Path(__file__).resolve().parents[2]  # scripts/ncm -> Chat UI
DEFAULT_EXCEL_PATH = _PROJECT_ROOT / "docs" / "NCM" / "Tabela_NCM.xlsx"
DEFAULT_OUTPUT_PATH = _PROJECT_ROOT / "docs" / "NCM" / "Tabela_NCM.json"

# Mapeamento de Levels e Capítulos
LEVELS_STRUCTURE = {
    "Level I": {
        "descricao": "Animais vivos e produtos do reino animal",
        "capitulos": {
            "01": "Animais vivos",
            "02": "Carnes e miudezas, comestíveis",
            "03": "Peixes e crustáceos, moluscos e os outros invertebrados aquáticos",
            "04": "Leite e laticínios; ovos de aves; mel natural; produtos comestíveis de origem animal, não especificados nem compreendidos em outros Capítulos",
            "05": "Outros produtos de origem animal, não especificados nem compreendidos em outros Capítulos"
        }
    },
    "Level II": {
        "descricao": "Produtos do reino vegetal",
        "capitulos": {
            "06": "Plantas vivas e produtos de floricultura",
            "07": "Produtos hortícolas, plantas, raízes e tubérculos, comestíveis",
            "08": "Frutas; cascas de cítricos e de melões",
            "09": "Café, chá, mate e especiarias",
            "10": "Cereais",
            "11": "Produtos da indústria de moagem; malte; amidos e féculas; inulina; glúten de trigo",
            "12": "Sementes e frutos oleaginosos; grãos, sementes e frutos diversos; plantas industriais ou medicinais; palha e forragem",
            "13": "Gomas, resinas e outros sucos e extratos vegetais",
            "14": "Matérias para entrançar e outros produtos de origem vegetal, não especificados nem compreendidos em outros Capítulos"
        }
    },
    "Level III": {
        "descricao": "Gorduras e óleos animais ou vegetais; produtos da sua dissociação; gorduras alimentares elaboradas; ceras de origem animal ou vegetal",
        "capitulos": {
            "15": "Gorduras e óleos animais ou vegetais; produtos da sua dissociação; gorduras alimentares elaboradas; ceras de origem animal ou vegetal"
        }
    },
    "Level IV": {
        "descricao": "Produtos das indústrias alimentares; bebidas, líquidos alcoólicos e vinagres; fumo (tabaco) e seus sucedâneos misturados",
        "capitulos": {
            "16": "Preparações de carne, de peixes ou de crustáceos, de moluscos ou de outros invertebrados aquáticos",
            "17": "Açúcares e produtos de confeitaria",
            "18": "Cacau e suas preparações",
            "19": "Preparações à base de cereais, farinhas, amidos, féculas ou de leite; produtos de pastelaria",
            "20": "Preparações de produtos hortícolas, de frutas ou de outras partes de plantas",
            "21": "Carnes e miudezas, comestíveis, salgadas ou em salmoura, secas ou defumadas (fumadas); farinhas e pós, comestíveis, de carnes ou de miudezas.",
            "22": "Bebidas, líquidos alcoólicos e vinagres",
            "23": "Resíduos e desperdícios das indústrias alimentares; alimentos preparados para animais",
            "24": "Fumo (tabaco) e seus sucedâneos, manufaturados"
        }
    },
    "Level V": {
        "descricao": "Produtos minerais",
        "capitulos": {
            "25": "Sal; enxofre; terras e pedras; gesso, cal e cimento",
            "26": "Minérios, escórias e cinzas",
            "27": "Combustíveis minerais, óleos minerais e produtos de sua destilação; matérias betuminosas; ceras minerais"
        }
    },
    "Level VI": {
        "descricao": "Produtos das indústrias químicas ou das indústrias conexas",
        "capitulos": {
            "28": "Produtos químicos inorgânicos; compostos inorgânicos ou orgânicos de metais preciosos, de elementos radioativos, de metais das terras raras ou de isótopos",
            "29": "Produtos químicos orgânicos",
            "30": "Produtos farmacêuticos",
            "31": "Adubos ou fertilizantes",
            "32": "Extratos tanantes e tintoriais; taninos e seus derivados; pigmentos e outras matérias corantes, tintas e vernizes, mástiques; tintas de escrever",
            "33": "Óleos essenciais e resinóides; produtos de perfumaria ou de toucador preparados e preparações cosméticas",
            "34": "Sabões, agentes orgânicos de superfície, preparações para lavagem, preparações lubrificantes, ceras artificiais, ceras preparadas, produtos de conservação e limpeza, velas e artigos semelhantes, massas ou pastas de modelar, ceras para dentistas e composições para dentista à base de gesso",
            "35": "Matérias albuminóides; produtos à base de amidos ou de féculas modificados; colas; enzimas",
            "36": "Pólvoras e explosivos; artigos de pirotecnia; fósforos; ligas pirofóricas; matérias inflamáveis",
            "37": "Produtos para fotografia e cinematografia",
            "38": "Produtos diversos das indústrias químicas"
        }
    },
    "Level VII": {
        "descricao": "Plásticos e suas obras; borracha e suas obras",
        "capitulos": {
            "39": "Plásticos e suas obras",
            "40": "Borracha e suas obras"
        }
    },
    "Level VIII": {
        "descricao": "Peles, couros, peleteria (peles com pêlo*) e obras desta matérias; artigos de correeiro ou de seleiro; artigos de viagem, bolsas e artefatos semelhantes; obras de tripa",
        "capitulos": {
            "41": "Peles, exceto a peleteria (peles com pêlo), e couros",
            "42": "Obras de couro; artigos de correeiro ou de seleiro; artigos de viagem, bolsas e artefatos semelhantes; obras de tripa",
            "43": "Peleteria (peles com pêlo) e suas obras; peleteria (peles com pêlo) artificial"
        }
    },
    "Level IX": {
        "descricao": "Madeira, carvão vegetal e obras de madeira; cortiça e suas obras; obras de espataria ou cestaria",
        "capitulos": {
            "44": "Madeira, carvão vegetal e obras de madeira",
            "45": "Cortiça e suas obras",
            "46": "Obras de espartaria ou de cestaria"
        }
    },
    "Level X": {
        "descricao": "Pastas de madeira ou de matérias fibrosas celulósicas; papel ou cartão de reciclar (desperdícios e aparas); papel e suas obras",
        "capitulos": {
            "47": "Pastas de madeira ou de outras matérias fibrosas celulósicas; papel ou cartão de reciclar (desperdícios e aparas)",
            "48": "Papel e cartão; obras de pasta de celulose, de papel ou de cartão",
            "49": "Livros, jornais, gravuras e outros produtos das indústrias gráficas; textos manuscritos ou datilografados, planos e plantas"
        }
    },
    "Level XI": {
        "descricao": "Matérias têxteis e suas obras",
        "capitulos": {
            "50": "Seda",
            "51": "Lã e pêlos finos ou grosseiros; fios e tecidos de crina",
            "52": "Algodão",
            "53": "Outras fibras têxteis vegetais; fios de papel e tecido de fios de papel",
            "54": "Filamentos sintéticos ou artificiais",
            "55": "Fibras sintéticas ou artificiais, descontínuas",
            "56": "Pastas (ouates), feltros e falsos tecidos; fios especiais; cordéis, cordas e cabos; artigos de cordoaria",
            "57": "Tapetes e outros revestimentos para pavimentos, de matérias têxteis",
            "58": "Tecidos especiais; tecidos tufados; rendas; tapeçarias; passamanarias; bordados",
            "59": "Tecidos impregnados, revestidos, recobertos ou estratificados; artigos para usos técnicos de matérias têxteis",
            "60": "Tecidos de malha",
            "61": "Vestuário e seus acessórios, de malha",
            "62": "Vestuário e seus acessórios, exceto de malha",
            "63": "Outros artefatos têxteis confeccionados; sortidos; artefatos de matérias têxteis, calçados, chapéus e artefatos de uso semelhante, usados; trapos"
        }
    },
    "Level XII": {
        "descricao": "Calçados, chapéus e artefatos de uso semelhante; guarda-chuvas, guarda-sóis, bengalas, chicotes, e suas partes; penas preparadas e suas obras; flores artificiais; obras de cabelo",
        "capitulos": {
            "64": "Calçados, polainas e artefatos semelhantes, e suas partes",
            "65": "Chapéus e artefatos de uso semelhante, e suas partes",
            "66": "Guarda-chuvas, sombrinhas, guarda-sóis, bengalas, bengalas-assentos, chicotes, e suas partes",
            "67": "Penas e penugem preparadas, e suas obras; flores arttificiais; obras de cabelo"
        }
    },
    "Level XIII": {
        "descricao": "Obras de pedra, gesso, cimento, amianto, mica ou de matérias semelhantes; produtos cerâmicos; vidro e suas obras",
        "capitulos": {
            "68": "Obras de pedra, gesso, cimento, amianto, mica ou de matérias semelhantes",
            "69": "Produtos cerâmicos",
            "70": "Vidro e suas obras"
        }
    },
    "Level XIV": {
        "descricao": "Pérolas naturais ou cultivadas, pedras preciosas ou semipreciosas e semelhantes; metais preciosos, metais folheados ou chapeados de metais preciosos, e suas obras; bijuterias; moedas",
        "capitulos": {
            "71": "Pérolas naturais ou cultivadas, pedras preciosas ou semipreciosas e semelhantes; metais preciosos, metais folheados ou chapeados de metais preciosos, e suas obras; bijuterias; moedas"
        }
    },
    "Level XV": {
        "descricao": "Metais comuns e suas obras",
        "capitulos": {
            "72": "Ferro fundido, ferro e aço",
            "73": "Obras de ferro fundido, ferro ou aço",
            "74": "Cobre e suas obras",
            "75": "Níquel e suas obras",
            "76": "Alumínio e suas obras",
            "78": "Chumbo e suas obras",
            "79": "Zinco e suas obras",
            "80": "Estanho e suas obras",
            "81": "Outros metais comuns; cermets; obras dessas matérias",
            "82": "Ferramentas, artefatos de cutelaria e talheres, e suas partes, de metais comuns",
            "83": "Obras diversas de metais comuns"
        }
    },
    "Level XVI": {
        "descricao": "Máquinas e aparelhos, material elétrico, e suas partes; aparelhos de gravação ou de reprodução de som, aparelhos de gravação ou de reprodução de imagens e de som em televisão, e suas partes e acessórios",
        "capitulos": {
            "84": "Reatores nucleares, caldeiras, máquinas, aparelhos e instrumentos mecânicos, e suas partes",
            "85": "Máquinas, aparelhos e materiais elétricos, e suas partes; aparelhos de gravação ou de reprodução de som, aparelhos de gravação ou de reprodução de imagens e de som em televisão, e suas partes e acessórios"
        }
    },
    "Level XVII": {
        "descricao": "Material de transporte",
        "capitulos": {
            "86": "Veículos e material para vias férreas ou semelhantes, e suas partes; aparelhos mecânicos (incluídos os eletromecânicos) de sinalização para vias de comunicação",
            "87": "Veículos automóveis, tratores, ciclos e outros veículos terrestres, suas partes e acessórios",
            "88": "Aeronaves e aparelhos espaciais, e suas partes",
            "89": "Embarcações e estruturas flutuantes"
        }
    },
    "Level XVIII": {
        "descricao": "Instrumentos e aparelhos de óptica, fotografia ou cinematografia, medida, controle ou de precisão; instrumentos e aparelhos médico-cirúrgicos; aparelhos de relojoaria; instrumentos musicais; suas partes e acessórios",
        "capitulos": {
            "90": "Instrumentos e aparelhos de óptica, fotografia ou cinematografia, medida, controle ou de precisão; instrumentos e aparelhos médico-cirúrgicos; suas partes e acessórios",
            "91": "Artigos de relojoaria",
            "92": "Instrumentos musicais, suas partes e acessórios"
        }
    },
    "Level XIX": {
        "descricao": "Armas e munições; suas partes e acessórios",
        "capitulos": {
            "93": "Armas e munições; suas partes e acessórios"
        }
    },
    "Level XX": {
        "descricao": "Mercadorias e produtos diversos",
        "capitulos": {
            "94": "Móveis, mobiliário médico-cirúrgico; colchões, almofadas e semelhantes; aparelhos de iluminação não especificados nem compreendidos em outros Capítulos; anúncios, cartazes ou tabuletas e placas indicadoras luminosos, e artigos semelhantes; construções pré-fabricadas",
            "95": "Brinquedos, jogos, artigos para divertimento ou para esporte; suas partes e acessórios",
            "96": "Obras diversas"
        }
    },
    "Level XXI": {
        "descricao": "Objetos de artes, de coleção e antiguidades",
        "capitulos": {
            "97": "Objetos de arte, de coleção e antiguidades"
        }
    },
    "Level 99": {
        "descricao": "(Reservado para usos especiais pelas Partes Contratantes)",
        "capitulos": {
            "999": "Códigos especiais (99900060, 99910000, etc.)"
        }
    }
}

def find_level_for_chapter(chapter_code):
    """Encontra o Level ao qual um capítulo pertence"""
    for level_name, level_data in LEVELS_STRUCTURE.items():
        if chapter_code in level_data["capitulos"]:
            return level_name
    return None

def convert_ncm_excel_to_json(excel_path, output_path=None):
    """
    Converte planilha Excel NCM para JSON estruturado
    
    Args:
        excel_path: Caminho para o arquivo Excel (.xlsx)
        output_path: Caminho para salvar o JSON (opcional, padrão: mesmo nome do Excel)
    """
    
    # Verificar se o arquivo existe
    excel_file = Path(excel_path)
    if not excel_file.exists():
        print(f"❌ Erro: Arquivo não encontrado: {excel_path}")
        return False
    
    print(f"📂 Carregando planilha: {excel_path}")
    
    try:
        # Carregar planilha
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        
        # Estrutura hierárquica baseada em Levels
        ncm_structure = {}
        
        # Inicializar todos os Levels
        for level_name, level_data in LEVELS_STRUCTURE.items():
            ncm_structure[level_name] = {
                "descricao": level_data["descricao"],
                "capitulos": {}
            }
            # Inicializar todos os capítulos do Level
            for chapter_code, chapter_desc in level_data["capitulos"].items():
                ncm_structure[level_name]["capitulos"][chapter_code] = {
                    "descricao": chapter_desc,
                    "ncms": {}
                }
        
        # Contadores
        total_rows = 0
        processed_rows = 0
        skipped_rows = 0
        ncm_by_level = {}  # Para estatísticas
        
        # Processar todas as linhas (começando da linha 1, sem cabeçalho)
        print("🔄 Processando linhas...")
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            total_rows += 1
            
            # Coluna A = código NCM, Coluna B = descrição
            codigo_raw = row[0] if len(row) > 0 else None
            descricao_raw = row[1] if len(row) > 1 else None
            
            # Validar dados
            if not codigo_raw or not descricao_raw:
                skipped_rows += 1
                continue
            
            # Converter para string e limpar
            codigo = str(codigo_raw).strip()
            descricao = str(descricao_raw).strip()
            
            # Remover caracteres não numéricos do código
            codigo_limpo = ''.join(filter(str.isdigit, codigo))
            
            # Validar código (deve ter pelo menos 2 dígitos para identificar capítulo)
            if not codigo_limpo or len(codigo_limpo) < 2:
                skipped_rows += 1
                continue
            
            # Validar descrição
            if not descricao or len(descricao) < 2:
                skipped_rows += 1
                continue
            
            # Extrair código do capítulo (2 primeiros dígitos)
            chapter_code = codigo_limpo[:2]
            
            # Encontrar o Level ao qual este capítulo pertence
            level_name = find_level_for_chapter(chapter_code)
            
            if not level_name:
                # Se não encontrou, verificar se é código especial (999...)
                if codigo_limpo.startswith("999"):
                    level_name = "Level 99"
                    chapter_code = "999"
                else:
                    skipped_rows += 1
                    continue
            
            # Adicionar NCM ao capítulo correspondente
            if level_name in ncm_structure:
                if chapter_code in ncm_structure[level_name]["capitulos"]:
                    ncm_structure[level_name]["capitulos"][chapter_code]["ncms"][codigo_limpo] = {
                        "codigo": codigo_limpo,
                        "descricao": descricao
                    }
                    processed_rows += 1
                    
                    # Estatísticas
                    if level_name not in ncm_by_level:
                        ncm_by_level[level_name] = {}
                    if chapter_code not in ncm_by_level[level_name]:
                        ncm_by_level[level_name][chapter_code] = 0
                    ncm_by_level[level_name][chapter_code] += 1
                else:
                    skipped_rows += 1
            else:
                skipped_rows += 1
        
        # Estatísticas
        print(f"\n✅ Processamento concluído!")
        print(f"   📊 Total de linhas: {total_rows}")
        print(f"   ✅ Processadas: {processed_rows}")
        print(f"   ⏭️  Ignoradas: {skipped_rows}")
        print(f"\n📈 Distribuição por Level e Capítulo:")
        for level_name in sorted(ncm_by_level.keys()):
            level_count = sum(ncm_by_level[level_name].values())
            print(f"   {level_name}: {level_count} NCMs")
            for chapter_code in sorted(ncm_by_level[level_name].keys()):
                count = ncm_by_level[level_name][chapter_code]
                print(f"      Capítulo {chapter_code}: {count} NCMs")
        
        # Determinar caminho de saída
        if output_path is None:
            # Se não especificado, usar mesmo nome do Excel na mesma pasta, mas com extensão .json
            output_path = excel_file.with_suffix('.json')
        else:
            output_path = Path(output_path)
        
        # Salvar JSON
        print(f"\n💾 Salvando JSON: {output_path}")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(ncm_structure, f, ensure_ascii=False, indent=2)
        
        # Estatísticas do arquivo JSON
        file_size = output_path.stat().st_size
        file_size_mb = file_size / (1024 * 1024)
        print(f"   ✅ Arquivo JSON salvo: {file_size_mb:.2f} MB")
        
        # Gerar também arquivo JS para evitar problemas de CORS
        js_output_path = output_path.with_suffix('.js')
        print(f"\n💾 Gerando arquivo JS: {js_output_path}")
        
        # Converter JSON para string e criar arquivo JS
        json_str = json.dumps(ncm_structure, ensure_ascii=False, indent=2)
        js_content = f"""// Dados NCM gerados automaticamente pelo convert_ncm_to_json.py
// Não edite este arquivo manualmente
// Última atualização: {datetime.now().isoformat()}

window.NCM_TABELA_DATA = {json_str};
"""
        
        with open(js_output_path, 'w', encoding='utf-8') as f:
            f.write(js_content)
        
        # Estatísticas do arquivo JS
        js_file_size = js_output_path.stat().st_size
        js_file_size_mb = js_file_size / (1024 * 1024)
        print(f"   ✅ Arquivo JS salvo: {js_file_size_mb:.2f} MB")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro ao processar: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Função principal"""
    print("=" * 60)
    print("🔄 CONVERSOR DE PLANILHA NCM PARA JSON")
    print("=" * 60)
    print()
    
    # Usar caminhos padrão ou argumentos fornecidos
    if len(sys.argv) >= 2:
        excel_path = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) > 2 else None
    else:
        # Usar caminhos padrão
        excel_path = DEFAULT_EXCEL_PATH
        output_path = DEFAULT_OUTPUT_PATH
        print(f"📂 Usando caminho padrão: {excel_path}")
        print(f"💾 Saída será salva em: {output_path}")
        print()
    
    # Converter
    success = convert_ncm_excel_to_json(excel_path, output_path)
    
    if success:
        print()
        print("=" * 60)
        print("✅ CONVERSÃO CONCLUÍDA COM SUCESSO!")
        print("=" * 60)
    else:
        print()
        print("=" * 60)
        print("❌ ERRO NA CONVERSÃO")
        print("=" * 60)
        sys.exit(1)


if __name__ == "__main__":
    main()

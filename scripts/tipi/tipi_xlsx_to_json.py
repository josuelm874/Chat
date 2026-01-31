#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor TIPI: planilha Excel ou CSV -> JSON (estrutura tipo Tabela_NCM)

Le tipi.xlsx ou tipi.csv com:
  - Coluna A: NCM
  - Coluna B: Reducao de Aliquota (%)
  - Coluna C: CST (3 digitos)
  - Coluna D: Classificacao Tributaria (6 digitos)

Gera JSON hierarquico Level -> capitulos -> ncms, no mesmo estilo de Tabela_NCM.json.

Uso: python tipi_xlsx_to_json.py [planilha.xlsx|.csv] [saida.json]
     python tipi_xlsx_to_json.py --sem-cabecalho tipi.xlsx   # planilha sem linha de cabecalho
     pip install openpyxl  # necessario apenas para .xlsx
"""

import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from _ncm_levels import LEVELS_STRUCTURE, find_level_for_chapter

def _load_openpyxl():
    try:
        from openpyxl import load_workbook
        return load_workbook
    except ImportError:
        return None

_PROJECT = Path(__file__).resolve().parents[2]
DEFAULT_EXCEL = _PROJECT / "docs" / "NCM" / "Tabela TIPI.xlsx"
DEFAULT_JSON = _PROJECT / "docs" / "NCM" / "Tabela_TIPI.json"


def _normalize_ncm(raw):
    s = str(raw or "").strip()
    digits = "".join(c for c in s if c.isdigit())
    return digits if len(digits) >= 2 else ""


def _normalize_reducao(raw):
    if raw is None or raw == "":
        return 0.0
    s = str(raw).strip().replace(",", ".")
    s = re.sub(r"%\s*$", "", s, flags=re.I)
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_cst(raw):
    s = str(raw or "").strip()
    digits = "".join(c for c in s if c.isdigit())
    return digits.zfill(3)[-3:]


def _normalize_classificacao(raw):
    s = str(raw or "").strip()
    digits = "".join(c for c in s if c.isdigit())
    return digits.zfill(6)[-6:]


def _normalize_header(txt):
    s = str(txt or "").strip().upper()
    for c, r in (("\u00c1", "A"), ("\u00c9", "E"), ("\u00cd", "I"), ("\u00d3", "O"), ("\u00da", "U"), ("\u00c7", "C"),
                  ("\u00c0", "A"), ("\u00c8", "E"), ("\u00cc", "I"), ("\u00d2", "O"), ("\u00d9", "U")):
        s = s.replace(c, r)
    return s


def _detect_columns(row):
    """Retorna (ix_ncm, ix_reducao, ix_cst, ix_class) ou None se usar A,B,C,D."""
    if not row or len(row) < 2:
        return (0, 1, 2, 3)
    hdrs = [_normalize_header(c) for c in row]
    ix_ncm = ix_reducao = ix_cst = ix_class = None
    ncm_kw = ("NCM", "CODIGO", "COD NCM", "CODIGO NCM")
    red_kw = ("REDUCAO", "ALIQUOTA", "REDUCAO DE ALIQUOTA")
    cst_kw = ("CST",)
    cls_kw = ("CLASSIFICACAO", "CLASS TRIB", "CLASSTRIB", "CLASS TRIBUTARIA", "CLASSTRIBUTARIA")
    for i, h in enumerate(hdrs):
        if not h:
            continue
        if ix_ncm is None and (h in ncm_kw or h.startswith("NCM") or ("CODIGO" in h and "NCM" in h)):
            ix_ncm = i
        if ix_reducao is None and (h in red_kw or "REDUCAO" in h or "ALIQUOTA" in h):
            ix_reducao = i
        if ix_cst is None and (h in cst_kw or h == "CST"):
            ix_cst = i
        if ix_class is None and (h in cls_kw or ("CLASS" in h and "TRIB" in h) or "CLASSTRIB" in h):
            ix_class = i
    if ix_ncm is not None and ix_reducao is not None and ix_cst is not None and ix_class is not None:
        return (ix_ncm, ix_reducao, ix_cst, ix_class)
    return (0, 1, 2, 3)


def _is_header_row(cells, col_map):
    if not cells or len(cells) < 2:
        return False
    ix = col_map[0]
    a = str(cells[ix] if ix < len(cells) else "" or "").strip().upper()
    return a in ("NCM", "CODIGO", "COD NCM", "CODIGO NCM") or "ncm" in (a[:10] if a else "")


def _iter_rows_xlsx(path, load_workbook_fn):
    wb = load_workbook_fn(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, values_only=True):
        yield list(row) if row else []
    wb.close()


def _iter_rows_csv(path):
    enc, sample = None, None
    for e in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            with open(path, "r", encoding=e, newline="") as f:
                sample = f.read(4096)
            enc = e
            break
        except UnicodeDecodeError:
            continue
    if enc is None:
        raise ValueError("Encoding nao suportado")
    first = (sample or "").split("\n")[0]
    delim = ";" if ";" in first else ","
    with open(path, "r", encoding=enc, newline="") as f:
        for row in csv.reader(f, delimiter=delim):
            yield row


def tipi_xlsx_to_json(input_path, output_path=None, sem_cabecalho=False):
    input_path = Path(input_path)
    if not input_path.exists():
        print("Erro: Arquivo nao encontrado:", input_path)
        return False

    print("Planilha:", input_path)
    if sem_cabecalho:
        print("Modo: sem cabecalho (A=NCM, B=Reducao, C=CST, D=Class.)")
    suf = input_path.suffix.lower()
    if suf == ".xlsx":
        load_wb = _load_openpyxl()
        if not load_wb:
            print("Erro: openpyxl nao instalado. Use: pip install openpyxl")
            return False
        rows = _iter_rows_xlsx(input_path, load_wb)
    elif suf == ".csv":
        try:
            rows = _iter_rows_csv(input_path)
        except Exception as e:
            print("Erro ao ler CSV:", e)
            return False
    else:
        print("Erro: Use .xlsx ou .csv")
        return False

    structure = {}
    for level_name, level_data in LEVELS_STRUCTURE.items():
        structure[level_name] = {"descricao": level_data["descricao"], "capitulos": {}}
        for ch, ch_desc in level_data["capitulos"].items():
            structure[level_name]["capitulos"][ch] = {"descricao": ch_desc, "ncms": {}}

    total = ok = skip = 0
    col_map = (0, 1, 2, 3)
    first = True
    preview = []

    for cells in rows:
        total += 1
        if first:
            first = False
            if not sem_cabecalho:
                col_map = _detect_columns(cells)
                if _is_header_row(cells, col_map):
                    skip += 1
                    continue
            else:
                col_map = (0, 1, 2, 3)
        i0, i1, i2, i3 = col_map
        a = cells[i0] if len(cells) > i0 else None
        b = cells[i1] if len(cells) > i1 else None
        c = cells[i2] if len(cells) > i2 else None
        d = cells[i3] if len(cells) > i3 else None
        codigo = _normalize_ncm(a)
        if not codigo:
            skip += 1
            continue
        reducao = _normalize_reducao(b)
        cst = _normalize_cst(c)
        classificacao = _normalize_classificacao(d)
        chapter = codigo[:2]
        if codigo.startswith("999"):
            chapter = "999"
        level_name = find_level_for_chapter(chapter)
        if not level_name:
            skip += 1
            continue
        cap = structure[level_name]["capitulos"].get(chapter)
        if not cap:
            skip += 1
            continue
        cap["ncms"][codigo] = {
            "codigo": codigo,
            "reducao_aliquota": round(reducao, 2),
            "cst": cst,
            "classificacao_tributaria": classificacao,
        }
        ok += 1
        if len(preview) < 5:
            preview.append((codigo, round(reducao, 2), cst, classificacao))

    print("Colunas usadas (NCM, Reducao, CST, Class.): indices", col_map)
    if preview:
        print("Preview (primeiras linhas da planilha no JSON):")
        for cod, red, cst, cls in preview:
            print("  ", cod, "| reducao:", red, "| cst:", cst, "| class:", cls)
    if total > 0 and (skip / total) > 0.5:
        print("Aviso: mais da metade das linhas foi ignorada. Verifique se A=NCM, B=Reducao, C=CST, D=Class. e se os capitulos NCM existem na tabela.")

    out = Path(output_path) if output_path else DEFAULT_JSON
    out.parent.mkdir(parents=True, exist_ok=True)
    print("JSON:", out)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(structure, f, ensure_ascii=False, indent=2)
    js_out = out.with_suffix(".js")
    js_content = "// TIPI – gerado por tipi_xlsx_to_json.py\n"
    js_content += "// " + datetime.now().isoformat() + "\n\n"
    js_content += "window.TIPI_TABELA_DATA = " + json.dumps(structure, ensure_ascii=False, indent=2) + ";\n"
    with open(js_out, "w", encoding="utf-8") as f:
        f.write(js_content)
    print("JS:  ", js_out)
    print("Total linhas:", total, "| Processadas:", ok, "| Ignoradas:", skip)
    return True


def main():
    args = [a for a in sys.argv[1:] if a != "--sem-cabecalho"]
    sem_cabecalho = "--sem-cabecalho" in sys.argv
    excel = args[0] if len(args) >= 1 else DEFAULT_EXCEL
    out = args[1] if len(args) >= 2 else None
    if not tipi_xlsx_to_json(excel, out, sem_cabecalho=sem_cabecalho):
        sys.exit(1)
    print("TIPI -> JSON concluido.")


if __name__ == "__main__":
    main()
